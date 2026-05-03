#!/usr/bin/env python3
"""
生成测试用 ZWF-20 打卡记录 Excel 文件
输出: /home/agentuser/shared/test_member_data.xlsx
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT_DIR = "/home/agentuser/shared"
OUT_FILE = os.path.join(OUT_DIR, "test_member_data.xlsx")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "打卡记录"

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell_align = Alignment(horizontal="center", vertical="center")

    # 表头（ZWF-20 设备导出格式）
    headers = [
        "会员ID", "姓名", "打卡时间", "验证方式", "进出方向",
        "门禁方式", "体温(℃)", "设备编号", "备注"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 模拟数据（10 条）
    import random
    members = [
        ("1001", "张三"),
        ("1002", "李四"),
        ("1003", "王五"),
        ("1004", "赵六"),
        ("1005", "孙七"),
    ]
    verify_modes = ["Face", "Card", "Card+Face", "Fingerprint"]
    in_outs = ["In", "In", "In", "Out"]  # 加权：上班多
    door_modes = ["hand_open", "auto", "hand_open"]

    records = [
        ("1001", "张三", "20260429083000", "Card+Face", "In", "hand_open", 36.5, "ZWF20-001"),
        ("1002", "李四", "20260429090000", "Face", "In", "hand_open", 36.7, "ZWF20-001"),
        ("1003", "王五", "20260429093000", "Card", "In", "auto", 36.3, "ZWF20-001"),
        ("1004", "赵六", "20260429100000", "Face", "In", "hand_open", 36.6, "ZWF20-001"),
        ("1005", "孙七", "20260429103000", "Card+Face", "In", "hand_open", 36.4, "ZWF20-001"),
        ("1001", "张三", "20260429120000", "Face", "Out", "hand_open", 36.5, "ZWF20-001"),
        ("1003", "王五", "20260429123000", "Face", "Out", "auto", 36.3, "ZWF20-001"),
        ("1002", "李四", "20260429180000", "Card", "Out", "hand_open", 36.7, "ZWF20-001"),
        ("1004", "赵六", "20260429183000", "Card+Face", "Out", "hand_open", 36.6, "ZWF20-001"),
        ("1005", "孙七", "20260429200000", "Face", "Out", "hand_open", 36.4, "ZWF20-001"),
    ]

    for row_idx, (mid, name, t, verify, io, door, temp, dev) in enumerate(records, 2):
        data = [
            mid, name, t, verify, io, door, temp, dev,
            f"{'上班' if io == 'In' else '下班'}打卡"
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # 列宽
    widths = [12, 10, 20, 14, 12, 12, 10, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(OUT_FILE)
    print(f"✅ 测试 Excel 已生成: {OUT_FILE}")
    print(f"   包含 {len(records)} 条模拟打卡记录")

    return OUT_FILE


if __name__ == "__main__":
    main()
